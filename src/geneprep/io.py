"""Input/output: FASTA/Excel readers with column autodetection and DNA input;
writers for optimised FASTA, QC report, and per-gene sequences workbook."""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

from Bio import SeqIO
from openpyxl import Workbook, load_workbook

from .codons import translate as translate_dna
from .qc import QCResult


@dataclass
class GeneEntry:
    name: str
    protein_sequence: str
    flank_5: str = ""
    flank_3: str = ""
    source_kind: str = "protein"    # 'protein' or 'dna' — how the user supplied it
    original_dna: str = ""          # if source_kind == 'dna', the input DNA verbatim


DNA_CHARS = set("ACGTNU")


def _looks_like_dna(seq: str) -> bool:
    if not seq:
        return False
    s = seq.upper().replace("*", "")
    dna = sum(1 for c in s if c in DNA_CHARS)
    return dna / max(len(s), 1) > 0.9 and len(s) >= 6


def _dna_to_protein(dna: str) -> str:
    """Translate DNA (frame 1). Trailing in-frame stop becomes '*'."""
    dna = dna.upper().replace("U", "T")
    dna = re.sub(r"[^ACGTN]", "", dna)
    # trim to multiple of 3
    dna = dna[:len(dna) - (len(dna) % 3)]
    aa = translate_dna(dna)
    return aa


# --- Column autodetection -----------------------------------------------

def _norm(s: str) -> str:
    return re.sub(r"[\s_'\"\-()]+", "", s.lower()) if s else ""


COLUMN_SYNONYMS: dict[str, list[str]] = {
    "name": ["name", "genename", "constructname", "id", "identifier", "label", "title", "construct"],
    "protein": ["proteinsequence", "protein", "aasequence", "aminoacidsequence",
                "aminoacid", "aa", "peptide", "peptidesequence", "aaseq", "proteinseq"],
    "dna": ["dnasequence", "dna", "nucleotidesequence", "nucleotide",
            "genesequence", "cds", "orf", "dnaseq", "nt", "ntsequence"],
    "flank5": ["5primeflank", "5flank", "fiveprimeflank", "5prime", "upstream",
               "forwardflank", "leftflank", "5end", "5", "5utr"],
    "flank3": ["3primeflank", "3flank", "threeprimeflank", "3prime", "downstream",
               "reverseflank", "rightflank", "3end", "3", "3utr"],
}


def detect_columns(headers: list[str]) -> dict[str, int | None]:
    """Map roles (name/protein/dna/flank5/flank3) to column indices."""
    normed = [_norm(h) for h in headers]

    def pick(role: str, exclude: set[int]) -> int | None:
        cands = COLUMN_SYNONYMS[role]
        # exact match first
        for cand in cands:
            for i, h in enumerate(normed):
                if i not in exclude and h == cand:
                    return i
        # substring match
        for cand in cands:
            for i, h in enumerate(normed):
                if i not in exclude and cand in h and h:
                    return i
        return None

    used: set[int] = set()
    mapping: dict[str, int | None] = {}
    for role in ("name", "protein", "dna", "flank5", "flank3"):
        idx = pick(role, used)
        mapping[role] = idx
        if idx is not None:
            used.add(idx)

    # If neither protein nor dna matched by name, try a generic 'sequence' column
    # and let content decide its type.
    if mapping["protein"] is None and mapping["dna"] is None:
        for i, h in enumerate(normed):
            if i in used:
                continue
            if "sequence" in h or "seq" in h:
                mapping["_unknown_seq"] = i
                used.add(i)
                break
    return mapping


def _first_nonempty(rows: list[list], col: int | None) -> str:
    if col is None:
        return ""
    for row in rows:
        if col < len(row) and row[col]:
            return str(row[col]).strip()
    return ""


def _confirm_mapping(headers: list[str], mapping: dict[str, int | None],
                     auto_yes: bool) -> None:
    """Print detected mapping and, if on a TTY and not --yes, confirm."""
    role_labels = {"name": "Name", "protein": "Protein sequence",
                   "dna": "DNA sequence", "flank5": "5' flank", "flank3": "3' flank",
                   "_unknown_seq": "Sequence (type auto-detected from content)"}
    print("  Column autodetection:")
    for role, idx in mapping.items():
        label = role_labels.get(role, role)
        if idx is None:
            continue
        print(f"    {label:35} <- column {idx + 1}: {headers[idx]!r}")
    if auto_yes or not sys.stdin.isatty():
        return
    ans = input("  Use this mapping? [Y/n] ").strip().lower()
    if ans not in ("", "y", "yes"):
        raise SystemExit(
            "Aborted. Rename your columns to match: "
            "name, protein_sequence (or dna_sequence), 5prime_flank, 3prime_flank."
        )


# --- Readers ------------------------------------------------------------

def read_fasta(path: Path) -> list[GeneEntry]:
    entries = []
    for record in SeqIO.parse(str(path), "fasta"):
        seq = str(record.seq).upper()
        if _looks_like_dna(seq):
            entries.append(GeneEntry(
                name=record.id, source_kind="dna", original_dna=seq,
                protein_sequence=_dna_to_protein(seq),
            ))
        else:
            entries.append(GeneEntry(
                name=record.id, source_kind="protein", protein_sequence=seq,
            ))
    return entries


def read_excel(path: Path, auto_yes: bool = False) -> list[GeneEntry]:
    wb = load_workbook(str(path), read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(c).strip() if c else "" for c in next(rows_iter)]
    data_rows = [list(r) for r in rows_iter]

    mapping = detect_columns(headers)
    if mapping.get("name") is None:
        raise SystemExit(
            f"Error: no 'name' column detected. Headers were: {headers}. "
            "Rename one column to 'name'."
        )
    if mapping.get("protein") is None and mapping.get("dna") is None \
            and "_unknown_seq" not in mapping:
        raise SystemExit(
            f"Error: no protein or DNA sequence column detected. "
            f"Headers were: {headers}. Rename one to 'protein_sequence' or 'dna_sequence'."
        )

    # Ambiguous 'sequence' column -> inspect content
    if "_unknown_seq" in mapping:
        idx = mapping["_unknown_seq"]
        sample = _first_nonempty(data_rows, idx)
        if _looks_like_dna(sample):
            mapping["dna"] = idx
        else:
            mapping["protein"] = idx
        del mapping["_unknown_seq"]

    _confirm_mapping(headers, mapping, auto_yes)

    name_c = mapping["name"]
    prot_c = mapping.get("protein")
    dna_c = mapping.get("dna")
    f5_c = mapping.get("flank5")
    f3_c = mapping.get("flank3")

    entries: list[GeneEntry] = []
    for row in data_rows:
        def cell(i):
            return str(row[i]).strip() if i is not None and i < len(row) and row[i] else ""
        name = cell(name_c)
        if not name:
            continue
        dna_val = cell(dna_c).upper() if dna_c is not None else ""
        prot_val = cell(prot_c).upper() if prot_c is not None else ""
        flank_5 = cell(f5_c).upper()
        flank_3 = cell(f3_c).upper()
        if dna_val and _looks_like_dna(dna_val):
            entries.append(GeneEntry(
                name=name, source_kind="dna", original_dna=dna_val,
                protein_sequence=_dna_to_protein(dna_val),
                flank_5=flank_5, flank_3=flank_3,
            ))
        elif prot_val:
            entries.append(GeneEntry(
                name=name, source_kind="protein",
                protein_sequence=prot_val,
                flank_5=flank_5, flank_3=flank_3,
            ))
    wb.close()
    return entries


def read_input(path: Path, auto_yes: bool = False) -> list[GeneEntry]:
    suffix = path.suffix.lower()
    if suffix in (".fa", ".fasta", ".faa", ".fas", ".fna"):
        return read_fasta(path)
    if suffix == ".xlsx":
        return read_excel(path, auto_yes=auto_yes)
    raise ValueError(f"Unsupported file format: {suffix}. Use .fasta or .xlsx")


# --- Writers ------------------------------------------------------------

def write_fasta(entries: list[tuple[str, str]], path: Path) -> None:
    with open(path, "w") as f:
        for name, seq in entries:
            f.write(f">{name}\n")
            for i in range(0, len(seq), 80):
                f.write(seq[i:i + 80] + "\n")


STATUS_COLORS = {
    "PASS": "C6EFCE",
    "REVIEW": "FFEB9C",
    "FAIL": "FFC7CE",
}


def write_optimised_xlsx(rows: list[dict], qc_results: list[QCResult],
                         path: Path) -> None:
    """Single workbook with two sheets:
      - 'Optimised': one row per gene, full construct in column 2 for easy copying
      - 'QC': one row per gene, all detector metrics
    """
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    seq_ws = wb.active
    seq_ws.title = "Optimised"
    seq_headers = ["Name", "Full DNA (order this)", "Protein (AA)",
                   "5' flank", "3' flank", "Optimised DNA (gene body)",
                   "Length (bp)", "Source of stop codon", "Status"]
    seq_ws.append(seq_headers)
    for cell in seq_ws[1]:
        cell.font = Font(bold=True)

    for r in rows:
        seq_ws.append([
            r["name"], r["full_dna"], r["protein"],
            r["flank_5"], r["flank_3"], r["optimised_dna"],
            len(r["full_dna"]), r.get("stop_source", ""), r["status"],
        ])
        sc = seq_ws.cell(row=seq_ws.max_row, column=seq_headers.index("Status") + 1)
        sc.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(r["status"], "FFEB9C"))
        sc.font = Font(bold=True)
        for col_i in (2, 3, 4, 5, 6):
            seq_ws.cell(row=seq_ws.max_row, column=col_i).alignment = Alignment(
                wrap_text=False, vertical="top")

    widths = {1: 32, 2: 95, 3: 60, 4: 22, 5: 22, 6: 80, 7: 12, 8: 22, 9: 10}
    for col_i, w in widths.items():
        seq_ws.column_dimensions[seq_ws.cell(row=1, column=col_i).column_letter].width = w
    seq_ws.freeze_panes = "C2"

    qc_ws = wb.create_sheet("QC")
    qc_headers = [
        "Gene Name", "Length (bp)", "GC%", "GC floor%", "CAI",
        "GC window min/max%", "Windows >85%", "Longest repeat (bp)",
        "Repeats >=15bp", "Max homopolymer", "5' GC%",
        "Restriction sites", "Source of stop codon", "Status", "Notes",
    ]
    qc_ws.append(qc_headers)
    for cell in qc_ws[1]:
        cell.font = Font(bold=True)

    for r in qc_results:
        qc_ws.append([
            r.name, r.length, r.gc_pct, r.gc_floor_pct, r.cai,
            f"{r.window_gc_min}-{r.window_gc_max}", r.windows_flagged,
            r.longest_repeat, r.repeats_flagged, r.max_homopolymer,
            r.five_prime_gc, r.restriction_str, r.stop_source,
            r.status, r.notes_str,
        ])
        sc = qc_ws.cell(row=qc_ws.max_row, column=qc_headers.index("Status") + 1)
        sc.fill = PatternFill("solid", fgColor=STATUS_COLORS.get(r.status, "FFEB9C"))
        sc.font = Font(bold=True)

    for col in qc_ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        qc_ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 55)
    qc_ws.freeze_panes = "A2"

    wb.save(str(path))
